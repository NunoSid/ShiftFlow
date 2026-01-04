import calendar
import io
import re
import zipfile
from collections import defaultdict
from typing import Dict, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from .config import TEMPLATE_PATH
from .constants import SHIFT_LOOKUP, WEEKDAY_EN, WEEKDAY_PT
from .models import (
    ConstraintEntry,
    ManualHoliday,
    Nurse,
    NurseMonthAdjustment,
    ProfessionalCategory,
    ScheduleEntry,
    Service,
    ServiceShift,
    SwapDecision,
    SwapParticipant,
    SwapRequest,
    User,
)
from .solver import collect_nurse_stats
from .utils import sort_nurses_by_category
from .holidays import month_holidays


EXPORT_LABELS = {
    "pt": {
        "schedule_title": "Escala",
        "availability_title": "Disponibilidades",
        "swaps_title": "Trocas",
        "professional": "Profissional",
        "previous_balance": "Débito anterior",
        "total_hours": "Total Horas",
        "current_balance": "Saldo atual",
        "monthly_hours": "Horas Mensais",
        "swap_date": "Data",
        "swap_requester": "Solicitante",
        "swap_participants": "Participantes",
        "swap_status": "Estado",
        "swap_current": "Atual (Serviço/Turno)",
        "swap_desired": "Pretendido (Serviço/Turno)",
        "swap_reason": "Motivo",
        "swap_observations": "Observações",
        "swap_decision": "Decisão",
        "swap_coordinator": "Coordenador",
        "swap_decision_date": "Data decisão",
    },
    "en": {
        "schedule_title": "Schedule",
        "availability_title": "Availability",
        "swaps_title": "Swaps",
        "professional": "Professional",
        "previous_balance": "Previous balance",
        "total_hours": "Total Hours",
        "current_balance": "Current balance",
        "monthly_hours": "Monthly Hours",
        "swap_date": "Date",
        "swap_requester": "Requester",
        "swap_participants": "Participants",
        "swap_status": "Status",
        "swap_current": "Current (Service/Shift)",
        "swap_desired": "Desired (Service/Shift)",
        "swap_reason": "Reason",
        "swap_observations": "Observations",
        "swap_decision": "Decision",
        "swap_coordinator": "Coordinator",
        "swap_decision_date": "Decision date",
    },
}

SWAP_STATUS_LABELS = {
    "pt": {
        "PENDING_PARTICIPANTS": "A aguardar participantes",
        "PENDING_COORDINATOR": "A aguardar coordenador",
        "APPROVED": "Aprovado",
        "REJECTED": "Rejeitado",
    },
    "en": {
        "PENDING_PARTICIPANTS": "Waiting for participants",
        "PENDING_COORDINATOR": "Waiting for coordinator",
        "APPROVED": "Approved",
        "REJECTED": "Rejected",
    },
}

PARTICIPANT_STATUS_LABELS = {
    "pt": {
        "PENDING": "A aguardar",
        "ACCEPTED": "Aceite",
        "REJECTED": "Recusado",
    },
    "en": {
        "PENDING": "Pending",
        "ACCEPTED": "Accepted",
        "REJECTED": "Rejected",
    },
}


def _lang_key(lang: str | None) -> str:
    return "en" if lang and lang.lower().startswith("en") else "pt"


def _label(lang: str | None, key: str, fallback: str | None = None) -> str:
    return EXPORT_LABELS[_lang_key(lang)].get(key, fallback or key)


def _day_headers(month: int, year: int, lang: str | None = None):
    _, days_in_month = calendar.monthrange(year, month)
    headers = []
    for day in range(1, days_in_month + 1):
        weekday_idx = calendar.weekday(year, month, day)
        weekday_names = WEEKDAY_EN if _lang_key(lang) == "en" else WEEKDAY_PT
        weekday = weekday_names[weekday_idx]
        headers.append((day, weekday))
    return headers


def _constraint_display(code: str) -> str:
    if not code:
        return ""
    mapping = {
        "FERIAS": "F",
        "DISPENSA": "DS",
        "PEDIDO_FOLGA": "P",
        "FERIADO": "FER",
        "FERIADO_TRAB": "FT",
        "INDISPONIVEL": "I",
        "DISPONIVEL": "D",
    }
    if code in mapping:
        return mapping[code]
    if code.startswith("DISPONIVEL_"):
        return f"D:{code.split('_', 1)[1]}"
    if code.startswith("INDISPONIVEL_"):
        return f"I:{code.split('_', 1)[1]}"
    return code[:4]


HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DAY_CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
SMALL_FONT = Font(size=8)
HOLIDAY_FILL = PatternFill(start_color="FFFAD1D5", end_color="FFFAD1D5", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="FFEFF4FF", end_color="FFEFF4FF", fill_type="solid")


def _weekend_set(month: int, year: int) -> set[int]:
    _, days_in_month = calendar.monthrange(year, month)
    weekends = set()
    for day in range(1, days_in_month + 1):
        weekday_idx = calendar.weekday(year, month, day)
        if weekday_idx in {5, 6}:
            weekends.add(day)
    return weekends


def _hex_to_fill(color: str) -> PatternFill:
    if not color:
        return None
    cleaned = color.lstrip("#")
    if len(cleaned) == 6:
        cleaned = f"FF{cleaned.upper()}"
    elif len(cleaned) == 8:
        cleaned = cleaned.upper()
    else:
        return None
    return PatternFill(start_color=cleaned, end_color=cleaned, fill_type="solid")


def _constraint_text(code: str) -> str:
    if not code:
        return ""
    mapping = {
        "FERIAS": "FERIAS",
        "DISPENSA": "DS",
        "FERIADO": "FER",
        "FERIADO_TRAB": "FT",
        "PEDIDO_FOLGA": "F",
        "PEDIDO_DESCANSO": "D",
        "PEDIDO_DESCANSO_FOLGA": "D/F",
        "INDISPONIVEL": "I",
        "DISPONIVEL": "V",
    }
    if code in mapping:
        return mapping[code]
    if code.startswith("DISPONIVEL_"):
        return code.split("_", 1)[1]
    if code.startswith("INDISPONIVEL_"):
        return f"I{code.split('_', 1)[1]}"
    return code[:4]


def _group_categories(session: Session, group: str | None) -> list[str]:
    if not group:
        return []
    normalized = group.strip().lower()
    role = None
    if normalized in {"ao", "assistente_operacional", "assistente operacional"}:
        role = "ASSISTENTE_OPERACIONAL"
    elif normalized in {"enf", "enfermagem"}:
        role = "ENFERMEIRO"
    else:
        role = session.scalar(
            select(ProfessionalCategory.role).where(
                ProfessionalCategory.role.ilike(group)
            )
        )
    if not role:
        return []
    return list(
        session.scalars(
            select(ProfessionalCategory.name).where(
                ProfessionalCategory.role == role,
                ProfessionalCategory.is_active.is_(True),
            )
        )
    )


def export_schedule(
    session: Session, year: int, month: int, group: str | None = None, lang: str | None = None
) -> io.BytesIO:
    workbook = _load_template_workbook()
    sheet_name = "SETEMBRO final 2025"
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    sheet["A1"] = (
        f"ShiftFlow - {_label(lang, 'schedule_title', 'Escala')} {month:02d}/{year}"
    )
    sheet["A3"] = _label(lang, "professional", "Profissional")
    sheet["B3"] = _label(lang, "previous_balance", "Débito anterior")
    sheet["A3"].font = SMALL_FONT
    sheet["B3"].font = SMALL_FONT

    day_headers = _day_headers(month, year, lang)
    weekend_set = _weekend_set(month, year)
    base_column = 3  # Column C
    max_days = 31

    manual_entries = [
        {
            "year": item.year,
            "month": item.month,
            "day": item.day,
            "label": item.label,
            "action": item.action,
        }
        for item in session.scalars(
            select(ManualHoliday).where(
                ManualHoliday.year == year,
                ManualHoliday.month == month,
            )
        )
    ]
    holiday_set = {record["day"] for record in month_holidays(year, month, manual_entries)}

    for idx in range(1, max_days + 1):
        column_letter = get_column_letter(base_column + idx - 1)
        if idx <= len(day_headers):
            day, weekday = day_headers[idx - 1]
            header_cell = sheet[f"{column_letter}3"]
            header_cell.value = f"{day}\n{weekday}"
            header_cell.alignment = HEADER_ALIGNMENT
            header_cell.font = SMALL_FONT
            if day in holiday_set:
                header_cell.fill = HOLIDAY_FILL
            elif day in weekend_set:
                header_cell.fill = WEEKEND_FILL
        else:
            sheet[f"{column_letter}3"] = ""

    summary_headers = [
        ("total_hours", _label(lang, "total_hours", "Total Horas")),
        ("current_balance", _label(lang, "current_balance", "Saldo atual")),
        ("monthly_hours", _label(lang, "monthly_hours", "Horas Mensais")),
    ]
    summary_columns = {}
    for offset, (key, label) in enumerate(summary_headers):
        column_idx = base_column + max_days + offset
        summary_columns[key] = column_idx
        summary_cell = sheet.cell(row=3, column=column_idx)
        summary_cell.value = label
        summary_cell.alignment = HEADER_ALIGNMENT
        summary_cell.font = SMALL_FONT

    categories = _group_categories(session, group)
    nurse_query = select(Nurse)
    if categories:
        nurse_query = nurse_query.where(Nurse.category.in_(categories))
    elif group:
        role = (group or "").strip()
        nurse_query = nurse_query.where(Nurse.category == role)
    nurses = sort_nurses_by_category(list(session.scalars(nurse_query)))
    nurse_ids = {nurse.id for nurse in nurses}
    if group and not nurse_ids:
        schedule_entries = []
    else:
        schedule_entries = list(
            session.scalars(
                select(ScheduleEntry).where(
                    ScheduleEntry.year == year,
                    ScheduleEntry.month == month,
                    ScheduleEntry.nurse_id.in_(nurse_ids) if nurse_ids else True,
                )
            )
        )
    day_assignments: Dict[Tuple[int, int], list[str]] = {}
    shift_colors = {
        row.shift_code: row.color
        for row in session.exec(
            select(ServiceShift.shift_code, Service.color)
            .join(Service, Service.code == ServiceShift.service_code)
        ).all()
    }
    assignment_minutes = defaultdict(int)
    for entry in schedule_entries:
        key = (entry.nurse_id, entry.day)
        day_assignments.setdefault(key, []).append(entry.shift_code)
        shift_meta = SHIFT_LOOKUP.get(entry.shift_code)
        if shift_meta:
            assignment_minutes[entry.nurse_id] += shift_meta.minutes

    if group and not nurse_ids:
        constraint_map = {}
    else:
        constraint_query = select(ConstraintEntry).where(
            ConstraintEntry.year == year,
            ConstraintEntry.month == month,
        )
        if nurse_ids:
            constraint_query = constraint_query.where(
                ConstraintEntry.nurse_id.in_(nurse_ids)
            )
        constraint_map: Dict[Tuple[int, int], str] = {
            (item.nurse_id, item.day): item.code
            for item in session.scalars(constraint_query)
        }

    if group and not nurse_ids:
        adjustments = {}
    else:
        adjustments_query = select(NurseMonthAdjustment).where(
            NurseMonthAdjustment.year == year,
            NurseMonthAdjustment.month == month,
        )
        if nurse_ids:
            adjustments_query = adjustments_query.where(
                NurseMonthAdjustment.nurse_id.in_(nurse_ids)
            )
        adjustments = {
            item.nurse_id: item for item in session.scalars(adjustments_query)
        }
    stats = collect_nurse_stats(session, nurses, year, month)
    stats_map = {item["nurse_id"]: item for item in stats}
    start_row = 4
    for row_offset, nurse in enumerate(nurses):
        row_idx = start_row + row_offset
        sheet.cell(row=row_idx, column=1).value = nurse.name
        stat = stats_map.get(nurse.id, {})
        previous_bank = stat.get("previous_bank_minutes", nurse.hour_balance_minutes or 0)
        sheet.cell(row=row_idx, column=2).value = round(previous_bank / 60, 2)

        for idx in range(1, max_days + 1):
            column_idx = base_column + idx - 1
            cell = sheet.cell(row=row_idx, column=column_idx)
            cell.alignment = DAY_CELL_ALIGNMENT
            cell.font = SMALL_FONT
            if idx in holiday_set:
                cell.fill = HOLIDAY_FILL
            elif idx in weekend_set:
                cell.fill = WEEKEND_FILL
            assignments = day_assignments.get((nurse.id, idx), [])
            if assignments:
                assignments_sorted = sorted(
                    assignments,
                    key=lambda code: (
                        SHIFT_LOOKUP.get(code).start_minute
                        if SHIFT_LOOKUP.get(code)
                        else 9999,
                        code,
                    ),
                )
                cell.value = "".join(assignments_sorted)
                if len(assignments_sorted) == 1:
                    color = shift_colors.get(assignments_sorted[0])
                    fill = _hex_to_fill(color) if color else None
                    if fill:
                        cell.fill = fill
                continue
            constraint = constraint_map.get((nurse.id, idx))
            if constraint:
                cell.value = _constraint_display(constraint)
            else:
                cell.value = ""

        adjustment = adjustments.get(nurse.id)
        actual_minutes = assignment_minutes.get(nurse.id, 0)
        if adjustment:
            actual_minutes += adjustment.extra_minutes
            actual_minutes -= adjustment.reduced_minutes
        target_minutes = stat.get("target_minutes", 0)
        saldo_atual_minutes = actual_minutes + previous_bank - target_minutes

        total_hours_cell = sheet.cell(row=row_idx, column=summary_columns["total_hours"])
        total_hours_cell.value = round(actual_minutes / 60, 2)
        total_hours_cell.font = SMALL_FONT
        saldo_cell = sheet.cell(row=row_idx, column=summary_columns["current_balance"])
        saldo_cell.value = round(saldo_atual_minutes / 60, 2)
        saldo_cell.font = SMALL_FONT
        target_cell = sheet.cell(row=row_idx, column=summary_columns["monthly_hours"])
        target_cell.value = round(target_minutes / 60, 2)
        target_cell.font = SMALL_FONT

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_constraints(
    session: Session, year: int, month: int, group: str | None = None, lang: str | None = None
) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _label(lang, "availability_title", "Disponibilidades")
    sheet["A1"] = _label(lang, "professional", "Profissional")

    day_headers = _day_headers(month, year, lang)
    header_alignment = HEADER_ALIGNMENT
    manual_entries = [
        {
            "year": item.year,
            "month": item.month,
            "day": item.day,
            "label": item.label,
            "action": item.action,
        }
        for item in session.scalars(
            select(ManualHoliday).where(
                ManualHoliday.year == year,
                ManualHoliday.month == month,
            )
        )
    ]
    holiday_set = {record["day"] for record in month_holidays(year, month, manual_entries)}
    for idx, (day, weekday) in enumerate(day_headers, start=2):
        cell = sheet.cell(row=1, column=idx)
        cell.value = f"{day}\n{weekday}"
        cell.alignment = header_alignment
        if day in holiday_set:
            cell.fill = HOLIDAY_FILL
        elif day in weekend_set:
            cell.fill = WEEKEND_FILL

    categories = _group_categories(session, group)
    nurse_query = select(Nurse)
    if categories:
        nurse_query = nurse_query.where(Nurse.category.in_(categories))
    nurses = sort_nurses_by_category(list(session.scalars(nurse_query)))
    nurse_ids = {nurse.id for nurse in nurses}
    if group and not nurse_ids:
        constraints_map = {}
    else:
        constraint_query = select(ConstraintEntry).where(
            ConstraintEntry.year == year,
            ConstraintEntry.month == month,
        )
        if nurse_ids:
            constraint_query = constraint_query.where(
                ConstraintEntry.nurse_id.in_(nurse_ids)
            )
        constraints_map: Dict[Tuple[int, int], str] = {
            (item.nurse_id, item.day): item.code
            for item in session.scalars(constraint_query)
        }

    for row_offset, nurse in enumerate(nurses, start=2):
        sheet.cell(row=row_offset, column=1).value = nurse.name
        for day_idx, (day, _) in enumerate(day_headers, start=1):
            column = day_idx + 1
            cell = sheet.cell(row=row_offset, column=column)
            cell.value = _constraint_text(constraints_map.get((nurse.id, day), ""))
            cell.alignment = DAY_CELL_ALIGNMENT
            if day in holiday_set:
                cell.fill = HOLIDAY_FILL
            elif day in weekend_set:
                cell.fill = WEEKEND_FILL

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_swaps(session: Session, lang: str | None = None) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _label(lang, "swaps_title", "Trocas")
    headers = [
        "ID",
        _label(lang, "swap_date", "Data"),
        _label(lang, "swap_requester", "Solicitante"),
        _label(lang, "swap_participants", "Participantes"),
        _label(lang, "swap_status", "Estado"),
        _label(lang, "swap_current", "Atual (Serviço/Turno)"),
        _label(lang, "swap_desired", "Pretendido (Serviço/Turno)"),
        _label(lang, "swap_reason", "Motivo"),
        _label(lang, "swap_observations", "Observações"),
        _label(lang, "swap_decision", "Decisão"),
        _label(lang, "swap_coordinator", "Coordenador"),
        _label(lang, "swap_decision_date", "Data decisão"),
    ]
    sheet.append(headers)

    users = {user.id: user for user in session.scalars(select(User))}
    decisions = {
        decision.request_id: decision
        for decision in session.scalars(select(SwapDecision))
    }
    participants_map: Dict[int, List[SwapParticipant]] = defaultdict(list)
    for participant in session.scalars(select(SwapParticipant)):
        participants_map[participant.request_id].append(participant)

    for swap in session.scalars(select(SwapRequest).order_by(SwapRequest.created_at.desc())):
        requester = users.get(swap.requester_id)
        participants = participants_map.get(swap.id, [])
        participants_label = []
        for part in participants:
            user = users.get(part.user_id)
            name = user.full_name if user else f"ID {part.user_id}"
            status_label = PARTICIPANT_STATUS_LABELS[_lang_key(lang)].get(
                part.status, part.status
            )
            participants_label.append(f"{name} ({status_label})")
        decision = decisions.get(swap.id)
        decision_text = (
            SWAP_STATUS_LABELS[_lang_key(lang)].get(decision.status, decision.status)
            if decision
            else ""
        )
        coordinator = users.get(decision.coordinator_id) if decision else None
        coordinator_name = coordinator.full_name if coordinator else ""
        decided_at = decision.decided_at.strftime("%d/%m/%Y %H:%M") if decision else ""
        current_shift = f"{swap.service_code or '-'} / {swap.shift_code or '-'}"
        desired_shift = f"{swap.desired_service_code or '-'} / {swap.desired_shift_code or '-'}"
        sheet.append(
            [
                swap.id,
                f"{swap.day:02d}/{swap.month:02d}/{swap.year}",
                requester.full_name if requester else swap.requester_id,
                " | ".join(participants_label),
                SWAP_STATUS_LABELS[_lang_key(lang)].get(swap.status, swap.status),
                current_shift,
                desired_shift,
                swap.reason or "",
                swap.observations or "",
                decision_text,
                coordinator_name,
                decided_at,
            ]
        )

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 22

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _load_template_workbook():
    try:
        return load_workbook(TEMPLATE_PATH)
    except TypeError as exc:
        if "workbookViewId" not in str(exc):
            raise
        with open(TEMPLATE_PATH, "rb") as fh:
            original = fh.read()
        buffer = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(original)) as source, zipfile.ZipFile(
            buffer, "w"
        ) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    text = data.decode("utf-8")
                    text = re.sub(r'workbookViewId="[^"]+"', "", text)
                    data = text.encode("utf-8")
                target.writestr(item, data)
        buffer.seek(0)
        return load_workbook(buffer)
